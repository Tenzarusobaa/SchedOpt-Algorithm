import mysql.connector
import pandas as pd
from openpyxl import load_workbook

# Database configuration
db_config = {
    'host': 'localhost',
    'user': 'root',
    'password': '', 
    'database': 'schedopt_db'
}

def import_rooms_from_excel(file_path):
    try:
        # Read the Excel file
        wb = load_workbook(file_path)
        sheet = wb['Filtered']
        
        # Get all rows from the sheet
        rows = list(sheet.iter_rows(values_only=True))
        
        # Find all the header rows (they contain "Room Code")
        header_indices = [i for i, row in enumerate(rows) if row and row[0] == "Room Code"]
        
        # Process each section between headers
        all_data = []
        for i in range(len(header_indices)):
            start = header_indices[i] + 1  # skip header row
            end = header_indices[i+1] if i+1 < len(header_indices) else len(rows)
            
            # Get the data rows for this section
            section_rows = rows[start:end]
            
            # Filter out empty rows and add to all_data
            for row in section_rows:
                if any(cell is not None and cell != '' for cell in row):
                    all_data.append(row)
        
        # Create DataFrame from the collected data
        columns = ["Room Code", "Building", "Capacity", "Size", "Type", "Function", 
                   "Department Owner", "Program Owner"]
        df = pd.DataFrame(all_data, columns=columns)
        
        # Clean data - replace empty strings with None and convert Capacity to float
        df = df.replace({'': None})
        df['Capacity'] = pd.to_numeric(df['Capacity'], errors='coerce')
        
        # Connect to MySQL
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor()
        
        # Prepare SQL query
        insert_query = """
        INSERT INTO tbl_room_data (
            rd_room_code, rd_building, rd_capacity, rd_size, 
            rd_type, rd_function, rd_department_owner, rd_program_owner
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """
        
        # Insert data row by row
        for _, row in df.iterrows():
            try:
                cursor.execute(insert_query, tuple(row))
            except mysql.connector.Error as err:
                print(f"Error inserting row {row}: {err}")
        
        # Commit changes and close connection
        conn.commit()
        cursor.close()
        conn.close()
        
        print(f"Successfully imported {len(df)} rooms into the database.")
        
    except Exception as e:
        print(f"An error occurred: {e}")

# Usage
if __name__ == "__main__":
    excel_file_path = "Rooms.xlsx"
    import_rooms_from_excel(excel_file_path)